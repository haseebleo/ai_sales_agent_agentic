"""
Lead Excel Writer — pandas-based, append-safe
Appends qualified leads to leads.xlsx (sheet: LeadData).

Why pandas instead of pure openpyxl:
- pandas writes to a NEW temp buffer first then atomically replaces the file
- this avoids file-lock conflicts when leads.xlsx is open in Microsoft Excel
- duplicate detection is done in-memory on the loaded DataFrame (fast, no row iteration)
- all 28 lead fields are supported

Thread safety: threading.Lock wraps all read-modify-write operations.
"""
from __future__ import annotations

import logging
import tempfile
import shutil
from datetime import datetime
from pathlib import Path
from threading import Lock
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.models import LeadData

logger = logging.getLogger("trango_agent.leads")

# All 28 columns in display order
LEAD_COLUMNS = [
    "lead_id",
    "session_id",
    "created_at",
    "updated_at",
    "source_channel",
    "full_name",
    "company_name",
    "email",
    "phone",
    "country",
    "industry",
    "team_size",
    "is_decision_maker",
    "interested_service",
    "recommended_package",
    "estimated_budget",
    "desired_timeline",
    "project_summary",
    "required_features",
    "preferred_platform",
    "lead_temperature",
    "lead_status",
    "payment_preference",
    "confidence_score",
    "retrieval_sources",
    "notes",
    "conversation_summary",
    "next_action",
]

HEADER_DISPLAY = {
    "lead_id": "Lead ID",
    "session_id": "Session ID",
    "created_at": "Created At",
    "updated_at": "Updated At",
    "source_channel": "Source Channel",
    "full_name": "Full Name",
    "company_name": "Company Name",
    "email": "Email",
    "phone": "Phone / WhatsApp",
    "country": "Country",
    "industry": "Industry",
    "team_size": "Team / Business Size",
    "is_decision_maker": "Decision Maker",
    "interested_service": "Interested Service",
    "recommended_package": "Recommended Package",
    "estimated_budget": "Estimated Budget",
    "desired_timeline": "Desired Timeline",
    "project_summary": "Project Summary",
    "required_features": "Required Features",
    "preferred_platform": "Platform (Web/Mobile/Both)",
    "lead_temperature": "Lead Temperature",
    "lead_status": "Lead Status",
    "payment_preference": "Payment Preference",
    "confidence_score": "Confidence Score",
    "retrieval_sources": "KB Sources Used",
    "notes": "Notes",
    "conversation_summary": "Conversation Summary",
    "next_action": "Next Action",
}

TEMP_COLORS = {
    "hot": "FF4C4C",
    "warm": "FFB347",
    "cold": "87CEEB",
}


def _lead_to_dict(lead: LeadData) -> dict:
    """Serialize a LeadData object to a flat dict matching LEAD_COLUMNS."""
    row = {}
    for col in LEAD_COLUMNS:
        val = getattr(lead, col, None)
        if val is None:
            row[col] = ""
        elif isinstance(val, datetime):
            row[col] = val.strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(val, bool):
            row[col] = "Yes" if val else "No"
        elif hasattr(val, "value"):
            row[col] = val.value  # Enum
        elif isinstance(val, float):
            row[col] = round(val, 3)
        else:
            row[col] = str(val) if val is not None else ""
    return row


def _apply_styling(file_path: Path, sheet_name: str) -> None:
    """
    Apply header formatting and temperature color-coding using openpyxl.
    Called AFTER pandas writes the base xlsx, so we don't fight pandas' engine.
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # ── Header row styling ────────────────────────────────────────────────────
    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Data rows: font + temperature color ───────────────────────────────────
    temp_col_idx = LEAD_COLUMNS.index("lead_temperature") + 1
    for row_num in range(2, ws.max_row + 1):
        temp_cell = ws.cell(row=row_num, column=temp_col_idx)
        temp_val = str(temp_cell.value or "").lower()
        temp_color = TEMP_COLORS.get(temp_val, "FFFFFF")

        for col_idx in range(1, len(LEAD_COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        temp_cell.fill = PatternFill("solid", start_color=temp_color)
        temp_cell.font = Font(name="Arial", size=9, bold=True)
        ws.row_dimensions[row_num].height = 16

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx in range(1, len(LEAD_COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    wb.save(file_path)


class LeadExcelWriter:
    """
    pandas-based lead writer.
    
    Strategy for Excel-open safety:
    1. Load existing file into a DataFrame (if it exists)
    2. Append the new row
    3. Write to a temp file using pandas to_excel()
    4. Atomic rename/replace of the temp file over the original
    
    Step 4 on macOS replaces atomically — even if Excel has the original open,
    Excel just re-reads on next access. The source file is never locked during write.
    """

    def __init__(self, file_path: str | None = None, sheet_name: str | None = None) -> None:
        self._file_path = Path(file_path or settings.LEADS_FILE_PATH)
        self._sheet_name = sheet_name or settings.LEAD_SHEET_NAME
        self._lock = Lock()
        self._file_path.parent.mkdir(parents=True, exist_ok=True)

    def _load_existing(self) -> pd.DataFrame:
        """Load existing leads into a DataFrame. Returns empty DataFrame if file absent."""
        if not self._file_path.exists():
            return pd.DataFrame(columns=LEAD_COLUMNS)
        try:
            df = pd.read_excel(
                self._file_path,
                sheet_name=self._sheet_name,
                dtype=str,
                engine="openpyxl",
            )
            # Normalize column headers back to field names
            reverse_header = {v: k for k, v in HEADER_DISPLAY.items()}
            df = df.rename(columns=reverse_header)
            # Keep only known columns in the right order
            for col in LEAD_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            return df[LEAD_COLUMNS].fillna("")
        except Exception as e:
            logger.warning(f"Could not load existing leads (will start fresh): {e}")
            return pd.DataFrame(columns=LEAD_COLUMNS)

    def _dup_key(self, email: str, phone: str, company: str) -> str:
        return f"{email.lower().strip()}|{phone.replace(' ','').strip()}|{company.lower().strip()}"

    def _existing_keys(self, df: pd.DataFrame) -> set[str]:
        keys = set()
        for _, row in df.iterrows():
            k = self._dup_key(
                str(row.get("email", "")),
                str(row.get("phone", "")),
                str(row.get("company_name", "")),
            )
            keys.add(k)
        return keys

    def _write_df(self, df: pd.DataFrame) -> None:
        """Write DataFrame to a temp file then atomically replace the target."""
        # Rename columns to display headers for the Excel file
        display_df = df.rename(columns=HEADER_DISPLAY)

        tmp_fd, tmp_path = tempfile.mkstemp(
            suffix=".xlsx",
            dir=self._file_path.parent,
            prefix=".tmp_leads_",
        )
        import os; os.close(tmp_fd)

        try:
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                display_df.to_excel(writer, sheet_name=self._sheet_name, index=False)

            _apply_styling(Path(tmp_path), self._sheet_name)
            shutil.move(tmp_path, self._file_path)
        except Exception:
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            raise

    def append_lead(self, lead: LeadData) -> bool:
        """
        Append lead to Excel. Returns True if written, False if duplicate.
        Thread-safe via internal Lock. Excel-safe via atomic temp-file replacement.
        """
        with self._lock:
            try:
                lead.updated_at = datetime.utcnow()
                df = self._load_existing()

                dup_key = self._dup_key(
                    str(lead.email or ""),
                    str(lead.phone or ""),
                    str(lead.company_name or ""),
                )
                existing_keys = self._existing_keys(df)

                if dup_key != "||" and dup_key in existing_keys:
                    logger.info(f"Duplicate lead skipped: {lead.lead_id} ({dup_key})")
                    return False

                new_row = _lead_to_dict(lead)
                new_df = pd.DataFrame([new_row])
                df = pd.concat([df, new_df], ignore_index=True)

                self._write_df(df)
                logger.info(
                    f"Lead written: {lead.lead_id} | "
                    f"{lead.lead_temperature.value} | "
                    f"total_rows={len(df)} | {self._file_path}"
                )
                return True

            except Exception as e:
                logger.error(f"Failed to write lead {lead.lead_id}: {e}", exc_info=True)
                raise

    def get_all_leads(self) -> list[dict]:
        """Read all leads back as list of dicts — for API /leads endpoint."""
        if not self._file_path.exists():
            return []
        try:
            df = pd.read_excel(
                self._file_path,
                sheet_name=self._sheet_name,
                dtype=str,
                engine="openpyxl",
            )
            df = df.fillna("")
            return df.to_dict(orient="records")
        except Exception as e:
            logger.warning(f"Could not read leads: {e}")
            return []
